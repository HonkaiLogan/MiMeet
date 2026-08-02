"""
从 xlsx 加载菜品和优惠数据，启动时缓存到内存
"""
import os
from pathlib import Path
import openpyxl

XLSX_PATH = Path(__file__).parent.parent / "database" / "科技园食堂菜品数据（2026-07-31晚餐）.xlsx"

_dishes: list[dict] = []
_offers: list[dict] = []


def _load():
    global _dishes, _offers
    if not XLSX_PATH.exists():
        print(f"[menu] xlsx not found: {XLSX_PATH}")
        return
    wb = openpyxl.load_workbook(str(XLSX_PATH))

    # 菜品 sheet
    ws1 = wb["食堂菜品"]
    _dishes = []
    for r in range(2, ws1.max_row + 1):
        row = [ws1.cell(r, c).value for c in range(1, 10)]
        if not row[0]:
            continue
        _dishes.append({
            "dish":      str(row[0] or ""),
            "price":     str(row[1] or ""),
            "unit":      str(row[2] or ""),
            "location":  str(row[3] or ""),
            "canteen":   str(row[4] or ""),
            "meal_time": str(row[5] or ""),
            "date":      str(row[6] or ""),
            "spicy":     str(row[7] or ""),
            "image":     str(row[8] or ""),
        })

    # 优惠 sheet
    ws2 = wb["羊毛优惠合集"]
    _offers = []
    for r in range(2, ws2.max_row + 1):
        row = [ws2.cell(r, c).value for c in range(1, 11)]
        if not row[0]:
            continue
        _offers.append({
            "merchant":    str(row[0] or ""),
            "category":    str(row[1] or ""),
            "area":        str(row[2] or ""),
            "address":     str(row[3] or ""),
            "phone":       str(row[4] or ""),
            "hours":       str(row[5] or ""),
            "discount":    str(row[6] or ""),
            "period":      str(row[7] or ""),
            "how_to_use":  str(row[8] or ""),
            "note":        str(row[9] or ""),
        })

    print(f"[menu] 加载完成：{len(_dishes)} 道菜品，{len(_offers)} 条优惠")


def get_dishes() -> list[dict]:
    return _dishes


def get_offers() -> list[dict]:
    return _offers


# 模块导入时自动加载
_load()
